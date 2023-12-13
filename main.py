import datetime
import json
import math
import os
import queue
import random
import re
import time
import traceback
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed

import bs4
import pandas as pd
import requests
from etaprogress.progress import ProgressBar
from furl import furl
from openpyxl.reader.excel import load_workbook
from requests.adapters import Retry, HTTPAdapter

proxy = []
cities = {
    "-7": "Актау",
    "-4": "Актобе",
    "-2": "Алматы",
    "-1": "Астана",
    "-8": "Атырау",
    "-37": "Балхаш",
    "-31": "Жанаозен",
    "-19": "Жезказган",
    "-3": "Караганда",
    "-9": "Кокшетау",
    "-21": "Конаев",
    "-10": "Костанай",
    "-11": "Кызылорда",
    "-5": "Павлодар",
    "-12": "Петропавловск",
    "-13": "Семей",
    "-14": "Талдыкорган",
    "-15": "Тараз",
    "-27": "Темиртау",
    "-29": "Туркестан",
    "-16": "Уральск",
    "-6": "Усть-Каменогорск",
    "-17": "Шымкент",
    "-18": "Экибастуз"
}
cities_en = {
    "-7": {"name": "Aktau", "code": "471010000"},
    "-4": {"name": "Aktobe", "code": "151010000"},
    "-2": {"name": "Almaty", "code": "750000000"},
    "-1": {"name": "Astana", "code": "710000000"},
    "-8": {"name": "Atyrau", "code": "231010000"},
    "-37": {"name": "Balkhash", "code": "351610000"},
    "-31": {"name": "Zhanaozen", "code": "471810000"},
    "-19": {"name": "Zhezkazgan", "code": "351810000"},
    "-3": {"name": "Karaganda", "code": "351010000"},
    "-9": {"name": "Kokshetau", "code": "111010000"},
    "-21": {"name": "Konaev", "code": "191610000"},
    "-10": {"name": "Kostanay", "code": "391010000"},
    "-11": {"name": "Kyzylorda", "code": "431010000"},
    "-5": {"name": "Pavlodar", "code": "551010000"},
    "-12": {"name": "Petropavlovsk", "code": "591010000"},
    "-13": {"name": "Semey", "code": "632810000"},
    "-14": {"name": "Taldykorgan", "code": "191010000"},
    "-15": {"name": "Taraz", "code": "311010000"},
    "-27": {"name": "Temirtau", "code": "352410000"},
    "-29": {"name": "Turkestan", "code": "512610000"},
    "-16": {"name": "Uralsk", "code": "271010000"},
    "-6": {"name": "Ust-Kamenogorsk", "code": "631010000"},
    "-17": {"name": "Shymkent", "code": "511010000"},
    "-18": {"name": "Ekibastuz", "code": "552210000"}
}

cats_eng = {
    '5': 'foto-i-video',
    '1': 'telefoni-i-gadzheti',
    '2': 'noutbuki-i-kompyuteri',
    '3': 'kuhonnaya-tehnika',
    '4': 'tehnika-dlya-doma',
    '6': 'televizori-i-audiotehnika',
    '7': 'avtotovari',
    '8': 'krasota-i-zdorove',
    '9': 'detskie-tovari',
    '10': 'naruchnye-chasy-sumki-i-aksessuary',
    '11': 'tovari-dlya-doma',
    '12': 'dosug',
    '14': 'tovari-dlya-zhivotnih',
    '31068': 'stroitelstvo-i-remont-',
    '32569': 'ukrashenija-aksessuary',
    '32626': 'mebel',
    '33119': 'dosug-i-tvorchestvo',
    '33819': 'apteka',
    '34319': 'specialnye-predlozhenija',
    '34672': 'odezhda',
    '34729': 'podarki-tovary-dlja-prazdnikov-i-cvety',
    '35891': 'halyk-shop',
    '37072': 'obuv',
    '38133': 'kanceljarskie-tovary',
    '39922': 'cifrovye-tovary'
}
cats_eng_swapped = {v: k for k, v in cats_eng.items()}
cats = {
    "5": "Фото и видео",
    "1": "Телефоны и гаджеты",
    "2": "Ноутбуки и компьютеры",
    "3": "Кухонная техника",
    "4": "Техника для дома",
    "6": "Телевизоры и аудиотехника",
    "7": "Автотовары",
    "8": "Красота и здоровье",
    "9": "Детские товары",
    "10": "Аксессуары",
    "11": "Товары для дома и дачи",
    "12": "Спорт и отдых",
    "14": "Товары для животных",
    "31068": "Строительство и ремонт",
    "32569": "Украшения",
    "32626": "Мебель",
    "33119": "Досуг и творчество",
    "33819": "Аптека",
    "34319": "Специальные предложения",
    "34672": "Одежда",
    "34729": "Подарки, товары для праздников и цветы",
    "35891": "Halyk shop",
    "37072": "Обувь",
    "38133": "Канцелярские товары",
    "39922": "Цифровые товары"
}

session = requests.Session()

retries = Retry(total=5,
                backoff_factor=0.1,
                status_forcelist=[500, 502, 503, 504])

session.mount('https://', HTTPAdapter(max_retries=retries, pool_maxsize=200))


def check_process(idn):
    r = requests.get("http://localhost:81/get_tasks")

    if not r.ok:
        return True
    result = int(idn) in r.json()

    print(f"[Check process]: ({idn} - {result})")
    return result


def clean_filename(name):
    forbidden_chars = '"*\\/\'|?:<>'
    filename = ''.join([x if x not in forbidden_chars else '#' for x in name])
    if len(filename) >= 176:
        filename = filename[:170] + '...'
    return filename


def proxy_checker(proxie_list):
    global proxy
    good = []
    for prox in proxie_list.split("\n"):
        if not prox.strip():
            return {}
        line = prox.strip().split(":")
        ip = line[0]
        port = line[1]
        if len(line) == 4:
            log = line[2]
            pwd = line[3]
            prox = {
                "http": f"socks5://{log}:{pwd}@{ip}:{port}",
                "https": f"socks5://{log}:{pwd}@{ip}:{port}",
            }
        else:
            prox = {
                "http": f"socks5://{ip}:{port}",
                "https": f"socks5://{ip}:{port}",
            }
        headers = {
            'authority': 'halykmarket.kz',
            'accept': 'application/json, text/plain, */*',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        }
        try:
            response = session.get("https://halykmarket.kz/", headers=headers, proxies=prox)
            if response.status_code == 200:
                print(f"[PROXY]: {prox} is good")
                good.append(prox)
            else:
                print(f"[PROXY]: {prox} isn't good")
        except:
            traceback.print_exc()
            continue
    proxy = good


def get_offers(sku, city, all_shops=None):
    headers = {
        "citycode": cities_en[city]['code'],
        'authority': 'halykmarket.kz',
        'accept': 'application/json, text/plain, */*',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }
    params = {
        'skuId': sku,
        'page': '1',
    }

    response = session.get('https://halykmarket.kz/api/public/v1/product/allMerchantOffersV2', params=params,
                           headers=headers, proxies=get_proxy()).json()
    if response.get("pagination"):
        if not isinstance(all_shops, list):
            return response['pagination']['totalNumberOfResults']

        else:
            all_shops += response['merchantInfoList']
            for page in range(1, response['pagination']['numberOfPages'] + 1):
                params.update({"page": str(page)})
                response = session.get('https://halykmarket.kz/api/public/v1/product/allMerchantOffersV2',
                                       params=params,
                                       headers=headers, proxies=get_proxy()).json()
                all_shops += response['merchantInfoList']
            return all_shops
    else:
        return None


def get_reviews(city, sku, n=10):
    if not n:
        return "", ""
    headers = {
        "citycode": cities_en[city]['code'],
        'authority': 'halykmarket.kz',
        'accept': 'application/json, text/plain, */*',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',

    }
    try:
        response = session.get(f'https://halykmarket.kz/api/public/v1/reviews/sku/{sku}', headers=headers,
                               proxies=get_proxy()).json()
    except:
        return get_reviews(city, sku, n - 1)
    total_reviews = response['allReviews']
    print(response)
    if total_reviews:
        last = datetime.datetime.fromtimestamp(
            round(response['reviewList'][0]['reviewSubmittedDate'] / 1000, 0)).strftime("%m.%d.%Y")
    else:
        last = ""
    return total_reviews, last


def get_checker_price(city, sku, n=10):
    if not n:
        return []
    headers = {
        'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
        'DNT': '1',
        'Accept-Language': 'ru',
        'sec-ch-ua-mobile': '?0',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36',
        'Accept': 'application/json, text/plain, */*',
        'Referer': 'https://halykmarket.kz/category/naushniki/naushniki-apple-airpods-pro-belye?recommended_by=dynamic&recommended_code=dcfba5d299eca404c51315645bb5336c',
        "citycode": cities_en[city]['code'],
        'sec-ch-ua-platform': '"Windows"',
    }

    params = {
        'skuId': sku,
        'page': '1',
        'installmentOrLoanPeriod': '0',
        'paymentType': 'CART',
        'sortBy': 'PRICE',
        'sortDirection': 'ASC',
    }
    shops = []
    while True:
        try:
            response = session.get('https://halykmarket.kz/api/public/v1/product/allMerchantOffersV2', params=params,
                                   headers=headers, proxies=get_proxy()).json()
        except:
            print("Error get_checker")
            return get_checker_price(city, sku, n - 1)
        if not response:
            break
        if response.get('merchantInfoList'):
            mInfo = response['merchantInfoList']
            for i, shop in enumerate(mInfo):
                disc = int(shop['loanInfo']["0"].get('discountPrice', 0)) or int(shop['price'])
                shops.append([shop['name'], int(shop['price']), disc, shop['id']])
        params['page'] = str(int(params['page']) + 1)
    return shops


def get_price(productUrl, city):
    try:
        imgs = [""] * 5
        shops = [["", "", "", ""] for _ in range(5)]
        headers = {
            "citycode": cities_en[city]['code'],
            'authority': 'halykmarket.kz',
            'accept': 'application/json, text/plain, */*',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',

        }
        params = {
            'productUrl': productUrl.split("?")[0],
        }
        response = session.post(
            'https://halykmarket.kz/api/public/v1/product/',
            json=params,
            headers=headers, proxies=get_proxy(),
        ).json()
        if isinstance(response, str):
            print(response)
            return None, None, None, None
        if response.get("imageUrls"):
            for i, img in enumerate(response['imageUrls'][:5]):
                imgs[i] = "https://cdn.halykmarket.kz" + img
        if response.get('merchantInfo'):
            mInfo = response['merchantInfo']
            for i, shop in enumerate(mInfo):
                disc = int(shop['loanInfo']["0"].get('discountPrice', 0)) or int(shop['price'])
                shops[i] = [shop['name'], int(shop['price']), disc, shop['id']]

        price = int(response.get("price") or 0)
        rating = response.get('rating', "")
        if rating:
            rating = rating['numberOfReviews']
        return price, imgs, rating, shops
    except:
        print("Error get_price")
        traceback.print_exc()


def get_review(link, n=10):
    if not n:
        return ""
    try:
        headers = {

            'authority': 'halykmarket.kz',
            'accept': 'application/json, text/plain, */*',
            'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
        }
        response = session.get(link, headers=headers, proxies=get_proxy())
        soup = bs4.BeautifulSoup(response.text, "lxml")
        scripts = soup.find_all("script")
        for i in scripts:
            if "window.__NUXT__=" in i.text:
                review = re.search("reviewSubmittedDate:(\d+)", i.text)
                if review:
                    review_time = datetime.date.fromtimestamp(int(review.group(1)) / 1000)
                    return review_time
                else:
                    return ""
    except requests.exceptions.ConnectionError:
        print("Error get_review")
        get_review(link, n - 1)


def product_thread(product, city, ind):
    print(f"Парсинг продукта: product({product}), ind({ind})")
    if not check_process(ind):
        raise Exception("Check process")
    try:
        link = "https://halykmarket.kz/category" + product['url']
        last_review = get_review(link)
        name = product['name']
        sku = product['id']
        price, imgs, rating, shops = get_price("/p" + product['url'], city)
        offers = get_offers(sku, city)
        main_cat = product['categories'][0]['name']
        categories = {i: cat.get('name', '') for i, cat in enumerate(product['categories'][:3])}
        price_in = product['price']
        price_out = product.get("oldprice", 0)
        data = {
            "Артикул": sku,
            "Бренд": product['brand'],
            "Название товара": name,
            "Рубрика": main_cat,
            "Категория №1": categories[1],
            "Категория №2": categories[2],
            "Фото1": imgs[0],
            "Фото2": imgs[1],
            "Фото3": imgs[2],
            "Фото4": imgs[3],
            "Отзывов": rating,
            "Размер скидки": product.get('discount', 0),
            "Цена со скидкой": price_in,
            "Цена без скидки": price_out,
            "Магазин1": shops[0][0],
            "Цена1": shops[0][1],
            "Магазин2": shops[1][0],
            "Цена2": shops[1][1],
            "Магазин3": shops[2][0],
            "Цена3": shops[2][1],
            "Магазин4": shops[3][0],
            "Цена4": shops[3][1],
            "Магазин5": shops[4][0],
            "Цена5": shops[4][1],
            "Продавцов": offers,
            "Последний отзыв": last_review,
            "Ссылка": link
        }
        return data
    except:
        print("Error product")
        traceback.print_exc()


def merchantInfo(merch):
    r = session.get(merch)
    soup = bs4.BeautifulSoup(r.text, "lxml")
    shop_name = soup.find('div', {'class': 'product-review-info-shop-block-texts-block-description'}).get_text(
        strip=True)
    shop_since = \
        soup.find('div', {'class': 'product-review-info-shop-block-texts-time-shop'}).get_text(strip=True).split(":")[
            1].strip()
    shop_phone = soup.find('a', {'class': 'product-review-info-shop-block-texts-number'}).get_text(strip=True)
    review_count = int(soup.find('div', {'class': 'product-review-rating-number'}).get_text(strip=True).split()[0])
    rating = float(soup.find('div', {'class': 'product-review-rating-title'}).get_text(strip=True))
    success_order = float(
        soup.find('div', {'class': 'product-review-success-order'}).find("div").get_text(strip=True).replace("%", ""))
    return {
        'Название Магазина': shop_name,
        'Дата регистрации': shop_since,
        'Телефон': shop_phone,
        'Количество отзывов': review_count,
        'Рейтинг': rating,
        'Успешных заказов': int(success_order)
    }


def get_proxy():
    if proxy:
        return random.choice(proxy)
    return None


def images_downloader(links):
    print(links)
    pass  # TODO: !


def get_products(mode, query, categories, city, wss: queue.Queue, proxie_list, ind, shop_checker=None, imgs=False,
                 bar=None,
                 params=None):
    if shop_checker is None:
        shop_checker = {}
    if not shop_checker:
        wss.put(
            json.dumps(
                {
                    "type": mode,
                    "curr": 0 if not shop_checker else shop_checker['curr'],
                    "total": 0 if not shop_checker else shop_checker['total'],
                    "eta": "-:-",
                    "id": ind,
                    "status": "Проверка прокси",
                    "name": query if not shop_checker else shop_checker['name'],
                    "city": cities[city],
                    "filename": ""
                }
            )
        )
        proxy_checker(proxie_list)
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }
    if not params:
        params = {
            'shop_id': '693ff081028570920fd8a6b971eb5e',
            'type': 'full_search',
            'search_query': query,
            'page': '1',
            'limit': '12',
            'filters': '{}',
            'sort_by': 'relevance',
            'order': 'asc',
            'categories': categories,
            'locations': city,
            'brands': '',
            'extended': 'true',
        }
    else:
        params.update({'search_query': query, 'categories': categories, 'locations': city})
    link_type = None
    brands = None
    mInfo = None
    kw = {}
    url = furl(query).args
    if url.get("f"):
        link_type = "filtered"
        url_split = url['f'].split(":")
        i = 0
        while i < len(url_split) - 1:
            kw[url_split[i]] = [url_split[i + 1]]
            i += 2
        brands = kw.get("brands", [])
        if brands:
            brands = brands[0]
        else:
            brands = ""
        kw.pop("brands", "")
        kw.pop("r46_search_query", "")

    if "r46_search_query" in query: query = url.get("r46_search_query")

    price_min = url.get("price_min", None)
    price_max = url.get("price_max", None)

    rgxp = re.search("/([a-zA-Z0-9\-]+)(/([a-zA-Z0-9\-]+))", query)
    if rgxp:
        categories = cats_eng_swapped.get(rgxp.group(3), "")
        query = cats.get(categories, query)
    else:
        query = kw.get("r46_search_query", query)
    params.update({"filters": json.dumps(kw) if kw else params['filters'], "brands": brands, "price_min": price_min,
                   "price_max": price_max,
                   "categories": categories, "search_query": query})
    filename = f"result_{cities[city]}_{query}_{mode}_{ind}.xlsx"
    if not shop_checker:
        wss.put(
            json.dumps(
                {
                    "type": mode,
                    "curr": 0,
                    "total": 0,
                    "eta": "-:-",
                    "id": ind,
                    "status": "Запуск",
                    "name": query,
                    "city": cities[city],
                    "filename": filename
                }
            )
        )
    response = session.get('https://api-r46.halykmarket.kz/search', params=params, headers=headers,
                           proxies=get_proxy())
    print("Parse products:", response)
    response = response.json()
    pages = math.ceil(response['products_total'] / 12)
    arr = []
    products = []
    if not shop_checker:
        bar = ProgressBar(pages, max_width=40)
    for page in range(1, pages + 1):
        if not check_process(ind):
            break
        if not shop_checker:
            bar.numerator = page

        params.update({"page": page})
        eta = str(bar).split("eta ")[1][:-2]
        retry = 5
        while retry:
            try:
                if not shop_checker:
                    wss.put(
                        json.dumps(
                            {
                                "type": mode,
                                "curr": page,
                                "total": pages,
                                "eta": eta,
                                "id": ind,
                                "status": "Сбор ссылок",
                                "name": query,
                                "city": cities[city]
                            }
                        )
                    )
                response = session.get('https://api-r46.halykmarket.kz/search', params=params, headers=headers,
                                       proxies=get_proxy()).json()
                for product in response['products']:
                    products.append(product)
                print(f"Сбор ссылок: retry({retry}), products({products})")
                break
            except KeyError:
                retry -= 1
    if not shop_checker:
        bar = ProgressBar(len(products), max_width=40)
    with ThreadPoolExecutor(max_workers=12) as executor:
        threads = []
        for product in products:
            threads.append(executor.submit(product_thread, product, city, ind))
        for result in as_completed(threads):
            print(f"Результат парсинга товара: result({result})")
            bar.numerator = len(arr) if not shop_checker else bar.numerator + 1
            res = result.result()
            if res:
                arr.append(result.result())
                eta = str(bar).split("eta ")[1][:-2]
                wss.put(
                    json.dumps(
                        {
                            "type": mode,
                            "curr": len(arr) if not shop_checker else bar.numerator,
                            "total": len(threads) if not shop_checker else bar.denominator,
                            "eta": eta,
                            "id": ind,
                            "status": "В работе",
                            "name": query if not shop_checker else shop_checker['name'],
                            "city": cities[city]
                        }
                    )
                )
            else:
                print("EXCPT")
    if link_type == "filtered":
        if kw.get("merchantName"):
            wss.put(
                json.dumps(
                    {
                        "type": mode,
                        "curr": len(arr),
                        "total": len(threads),
                        "eta": "-:-",
                        "id": ind,
                        "status": "Обработка магазина",
                        "name": query,
                        "city": cities[city]
                    }
                )
            )
            for shop in arr:
                for merchantName in get_offers(shop['Артикул'], city, []):
                    if merchantName['name'] == kw['merchantName'][0]:
                        mInfo = merchantInfo(f"https://halykmarket.kz/reviews/merchant/{merchantName['id']}")
                        break
                if mInfo:
                    break

    if not mode == "shop":
        if imgs:
            total = 0
            for line in arr:
                for key, value in line.items():
                    if key.startswith('Фото') and value:
                        total += 1
            wss.put(
                json.dumps(
                    {
                        "type": mode,
                        "curr": 0,
                        "total": total,
                        "eta": "-:-",
                        "id": ind,
                        "status": "Скачивание картинок ",
                        "name": query,
                        "city": cities[city]
                    }
                )
            )
            bar = ProgressBar(total, max_width=40)
            for line in arr:
                result_dir = clean_filename(filename.replace(".xlsx", ""))
                if not os.path.isdir(result_dir):
                    os.mkdir(result_dir)
                item_dir = os.path.join(result_dir, clean_filename(line['Название товара']))
                if not os.path.isdir(item_dir):
                    os.mkdir(item_dir)
                for key, value in line.items():
                    if key.startswith('Фото') and value:
                        eta = str(bar).split("eta ")[1][:-2]
                        img_url = value
                        img_ext = os.path.splitext(img_url)[1]
                        img_filename = clean_filename(f"{line['Название товара']}_{key}{img_ext}")
                        img_path = os.path.join(item_dir, img_filename)
                        retry = 10
                        while retry:
                            if not check_process(ind):
                                break
                            try:
                                rr = requests.get(img_url, proxies=get_proxy())
                                if rr.status_code == 404:
                                    break
                                rr.raise_for_status()
                                with open(img_path, 'wb') as f:
                                    f.write(rr.content)
                                break
                            except (requests.exceptions.RequestException, OSError) as e:
                                print(f"Failed to download image {img_url}, retrying", e)
                                time.sleep(1)
                                retry -= 1
                        bar.numerator += 1
                        wss.put(
                            json.dumps(
                                {
                                    "type": mode,
                                    "curr": bar.numerator,
                                    "total": total,
                                    "eta": eta,
                                    "id": ind,
                                    "status": "Скачивание картинок",
                                    "name": query,
                                    "city": cities[city]
                                }
                            )
                        )
        print(f"Выгрузка файла СТАРТ")
        df = pd.DataFrame(arr)
        writer = pd.ExcelWriter(filename, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_urls': False}})
        df.to_excel(writer, sheet_name='Sheet1', index=False)
        convert = lambda x: x * 4.91803278689
        worksheet = writer.sheets['Sheet1']
        worksheet.set_column(0, 0, convert(2.37))
        worksheet.set_column(1, 1, convert(4.5))
        worksheet.set_column(2, 4, convert(3))
        worksheet.set_column(9, 9, convert(1.8))
        worksheet.set_column(10, 11, convert(3.15))
        worksheet.set_column(12, 12, convert(4.72))
        worksheet.set_column(13, 13, convert(2))
        worksheet.set_column(14, 14, convert(4.72))
        worksheet.set_column(15, 15, convert(2))
        worksheet.set_column(16, 16, convert(4.72))
        worksheet.set_column(17, 17, convert(2))
        worksheet.set_column(18, 18, convert(4.72))
        worksheet.set_column(19, 19, convert(2))
        worksheet.set_column(20, 20, convert(4.72))
        worksheet.set_column(21, 21, convert(2))
        worksheet.set_column(22, 22, convert(2.24))
        worksheet.set_column(23, 23, convert(3.36))
        writer.close()

        if kw.get("merchantName"):
            wb = load_workbook(filename=filename)
            worksheet = wb.active
            rows = ["Название Магазина", "Дата регистрации", "Рейтинг", "Телефон",
                    "Количество отзывов", "Успешных заказов"]
            worksheet.insert_rows(0, 9)
            for i, row in enumerate(rows, start=2):
                worksheet[f"B{i}"] = row
                worksheet[f"C{i}"] = mInfo[row]
            wb.save(filename=filename)
        if imgs:
            wss.put(
                json.dumps(
                    {
                        "type": mode,
                        "curr": bar.numerator,
                        "total": bar.numerator,
                        "eta": "-:-",
                        "id": ind,
                        "status": "Создание файла",
                        "name": query,
                        "city": cities[city]
                    }
                )
            )
            with zipfile.ZipFile(filename.replace("xlsx", "zip"), mode="w") as archive:
                for dirname, subdirs, files in os.walk(filename.replace(".xlsx", "")):
                    archive.write(dirname)
                    for filename_dir in files:
                        archive.write(os.path.join(dirname, filename_dir))
                archive.write(filename)
                filename = filename.replace("xlsx", "zip")
        wss.put(
            json.dumps(
                {
                    "type": mode,
                    "curr": len(threads),
                    "total": len(threads),
                    "eta": "-:-",
                    "id": ind,
                    "status": "Готово",
                    "name": query,
                    "city": cities[city],
                    "filename": filename
                }
            )
        )
        print(f"Выгрузка файла КОНЕЦ")
    else:
        return arr


def search_by_name(name, ind):
    headers = {
        'Accept': '*/*',
        'Accept-Language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7,zh-TW;q=0.6,zh;q=0.5,no;q=0.4,uk;q=0.3',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Content-Type': 'application/x-www-form-urlencoded',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36',
        'sec-ch-ua': '"Chromium";v="112", "Google Chrome";v="112", "Not:A-Brand";v="99"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }
    params = {
        'shop_id': '693ff081028570920fd8a6b971eb5e',
        'did': 'vuoHqksmAD',
        'sid': '1PMrRSuTyF',
        'type': 'instant_search',
        'search_query': re.sub("бренд (.*)", "", f'{name}'),
        'referer': 'https://halykmarket.kz/',
    }
    discount = 0
    try:
        response = requests.get('https://api-r46.halykmarket.kz/search', params=params, headers=headers,
                                proxies=get_proxy()).json()['products']
        url = ""
        sku = 0

        if response:
            sku = response[0]['id']
            url = response[0]['url']
            name = response[0]['name']
            discount = response[0].get('discount', 0)
        else:
            name = ""
    except Exception as e:
        print(f"Error search_by_name, {e}")
        return search_by_name(name, ind)
    return sku, url, name, discount


def checker_thread(line, city, index, ind):
    if not check_process(ind):
        raise Exception("Check process")
    try:
        # link_price = "/".join(line['Ссылка'].split("/")[4:])
        sku, link, name, discount = search_by_name(line['Name'], ind)
        if not sku:
            return {
                "review": "",
                "flagman_price": "",
                "goldmoon_price": "",
                "flagman_index": "",
                "goldmoon_index": "",
                "price": "",
                "price_old": "",
                "rating": "",
                "sku": line['SKU'],
                "sku_h": sku,
                "index": index,
                "link": link,
                "sellers": 0,
                "name": name
            }
        shops = get_checker_price(city, sku)
        rating, review = get_reviews(city, sku)
        prices_old = [x[1] for x in shops if x[1]]
        prices = [x[2] for x in shops if x[2]]
        price_old = min(prices_old, default="")
        price = min(prices, default="")
        shops_names = [x[0] for x in shops]
        flagman_index = shops_names.index("Flagman") + 1 if "Flagman" in shops_names else ""
        flagman_price = shops[flagman_index - 1][2] if isinstance(flagman_index, int) else ""
        goldmoon_index = shops_names.index("GoldMoon") + 1 if "GoldMoon" in shops_names else ""
        goldmoon_price = shops[goldmoon_index - 1][2] if isinstance(goldmoon_index, int) else ""
        return {
            "review": review,
            "flagman_price": flagman_price,
            "goldmoon_price": goldmoon_price,
            "flagman_index": flagman_index,
            "goldmoon_index": goldmoon_index,
            "discount": discount,
            "price": price,
            "price_old": price_old,
            "rating": rating,
            "sku": line['SKU'],
            "sku_h": sku,
            "index": index,
            "link": "https://halykmarket.kz/category" + link,
            "sellers": len(shops),
            "name": name
        }
    except:
        print("Error checker_thread")
        traceback.print_exc()


def checker(mode, lines, prox, city, ind, ignore, filename, wss: queue.Queue):
    filename = filename.replace(".xlsx", "_result.xlsx")
    print(filename)
    wss.put(json.dumps(
        {"type": mode, "curr": 0, "total": 0, "eta": "-:-", "id": ind, "status": "Проверка прокси",
         "name": filename, "city": cities[city]}))
    proxy_checker(prox)
    wss.put(json.dumps(
        {"type": mode, "curr": 0, "total": 0, "eta": "-:-", "id": ind, "status": "Запуск",
         "name": filename, "city": cities[city]}))
    with ThreadPoolExecutor(max_workers=10) as executor:
        threads = []
        for index, line in enumerate(lines):
            if line['SKU'] in ignore:
                continue
            threads.append(executor.submit(checker_thread, line, city, index, ind))
        bar = ProgressBar(len(threads))
        for result in as_completed(threads):
            bar.numerator += 1
            result = result.result()
            if not result:
                continue
            row = result['index']
            lines[row]['Название Halyk'] = result['name']
            lines[row]['Кол-во продавцов'] = result['sellers']
            lines[row]['Ссылка на товар Halyk'] = result['link']
            lines[row]['SKU Halyk'] = result['sku_h']
            lines[row]['Размер скидки'] = result.get('discount')
            lines[row]['Цена со скидкой'] = result['price']
            lines[row]['Цена без скидки'] = result['price_old']
            lines[row]['Цена Flagman'] = result['flagman_price']
            lines[row]['Место №'] = result['flagman_index']
            lines[row]['Цена GoldMoon'] = result['goldmoon_price']
            lines[row]['Место №.1'] = result['goldmoon_index']
            lines[row]['Последний отзыв'] = result['review']
            lines[row]['Кол-во отзывов'] = result['rating']
            eta = str(bar).split("eta ")[1][:-2]

            wss.put(json.dumps(
                {"type": mode, "curr": bar.numerator, "total": len(threads), "eta": eta, "id": ind,
                 "status": "В работе",
                 "name": filename, "city": cities[city]}))
    df = pd.DataFrame(lines)
    writer = pd.ExcelWriter(f"/root/Halyk/static/{filename}", engine='xlsxwriter',
                            engine_kwargs={'options': {'strings_to_urls': False}})
    df.to_excel(writer, sheet_name='Sheet1', index=False)

    convert = lambda x: x * 4.91803278689
    worksheet = writer.sheets['Sheet1']
    worksheet.set_column(0, 0, convert(2.5))
    worksheet.set_column(11, 11, convert(3.2))
    writer.close()
    wss.put(json.dumps({
        "type": "checker",
        "curr": len(threads),
        "total": len(threads),
        "eta": "-:-",
        "id": ind,
        "status": "Готово",
        "name": filename,
        "city": cities[city],
        "filename": filename
    }))


def store_parser(mode, query, categories, city, wss: queue.Queue, proxie_list, ind, store_mode, pictures):
    wss.put(
        json.dumps(
            {
                "type": mode,
                "curr": 0,
                "total": 0,
                "eta": "-:-",
                "id": ind,
                "status": "Проверка прокси",
                "name": query,
                "city": cities[city],
                "filename": ''
            }
        )
    )
    proxy_checker(proxie_list)
    headers = {
        'Accept': 'application/json, text/plain, */*',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36',
    }
    merch = merchantInfo(query)
    query = merch["Название Магазина"]
    filename = f"{cities[city]}_{query}.xlsx"
    params = {
        'shop_id': '693ff081028570920fd8a6b971eb5e',
        'type': 'full_search',
        'page': '1',
        'limit': '12',
        'filters': json.dumps({'merchantName': [merch["Название Магазина"]]}),
        'sort_by': 'relevance',
        'order': 'asc',
        'categories': categories,
        'locations': city,
        'brands': '',
        'extended': 'true',
    }
    results = []
    total = 0
    wss.put(
        json.dumps(
            {
                "type": mode,
                "curr": 0,
                "total": 0,
                "eta": "-:-",
                "id": ind,
                "status": "Сбор данных",
                "name": query,
                "city": cities[city],
                "filename": filename
            }
        )
    )
    for k, cat in cats_eng.items():
        params.update({"categories": k, "search_query": cats[k]})
        response = session.get('https://api-r46.halykmarket.kz/search', params=params, headers=headers,
                               proxies=get_proxy())
        print(response.url)
        response = response.json()
        total += response['products_total']
        if not check_process(ind):
            break
    bar = ProgressBar(total)
    wss.put(
        json.dumps(
            {
                "type": mode,
                "curr": 0,
                "total": total,
                "eta": "-:-",
                "id": ind,
                "status": "В работе",
                "name": query,
                "city": cities[city],
                "filename": filename
            }
        )
    )

    for k, cat in cats_eng.items():
        print(bar, end='\r')
        merch_data = {"name": merch["Название Магазина"], "curr": bar.numerator, "total": total}
        results += get_products(mode, cats[k], categories, city, wss, proxie_list, ind, merch_data, pictures, bar,
                                params)
    if pictures:
        total = 0
        for line in results:
            for key, value in line.items():
                if key.startswith('Фото') and value:
                    total += 1
        wss.put(
            json.dumps(
                {
                    "type": mode,
                    "curr": 0,
                    "total": total,
                    "eta": "-:-",
                    "id": ind,
                    "status": "Скачивание картинок ",
                    "name": query,
                    "city": cities[city]
                }
            )
        )
        bar = ProgressBar(total, max_width=40)
        for line in results:
            result_dir = clean_filename(filename.replace(".xlsx", ""))
            if not os.path.isdir(result_dir):
                os.mkdir(result_dir)
            item_dir = os.path.join(result_dir, clean_filename(line['Название товара']))
            if not os.path.isdir(item_dir):
                os.mkdir(item_dir)
            for key, value in line.items():
                if key.startswith('Фото') and value:
                    eta = str(bar).split("eta ")[1][:-2]
                    img_url = value
                    img_ext = os.path.splitext(img_url)[1]
                    img_filename = clean_filename(f"{line['Название товара']}_{key}{img_ext}")
                    img_path = os.path.join(item_dir, img_filename)
                    retry = 10
                    while retry:
                        try:
                            rr = requests.get(img_url, proxies=get_proxy())
                            if rr.status_code == 404:
                                break
                            rr.raise_for_status()
                            with open(img_path, 'wb') as f:
                                f.write(rr.content)
                            break
                        except (requests.exceptions.RequestException, OSError) as e:
                            print(f"Failed to download image {img_url}, retrying", e)
                            time.sleep(1)
                            retry -= 1
                    bar.numerator += 1
                    wss.put(
                        json.dumps(
                            {
                                "type": mode,
                                "curr": bar.numerator,
                                "total": total,
                                "eta": eta,
                                "id": ind,
                                "status": "Скачивание картинок",
                                "name": query,
                                "city": cities[city]
                            }
                        )
                    )
    if pictures:
        wss.put(
            json.dumps(
                {
                    "type": mode,
                    "curr": bar.numerator,
                    "total": bar.numerator,
                    "eta": "-:-",
                    "id": ind,
                    "status": "Создание файла",
                    "name": query,
                    "city": cities[city]
                }
            )
        )

    df = pd.DataFrame(results)
    writer = pd.ExcelWriter(filename, engine='xlsxwriter', engine_kwargs={'options': {'strings_to_urls': False}})

    df.to_excel(writer, sheet_name='Sheet1', index=False)
    convert = lambda x: x * 4.91803278689
    worksheet = writer.sheets['Sheet1']
    worksheet.set_column(0, 0, convert(2.37))
    worksheet.set_column(1, 1, convert(4.5))
    worksheet.set_column(2, 4, convert(3))
    worksheet.set_column(9, 9, convert(1.8))
    worksheet.set_column(10, 11, convert(3.15))
    worksheet.set_column(12, 12, convert(4.72))
    worksheet.set_column(13, 13, convert(2))
    worksheet.set_column(14, 14, convert(4.72))
    worksheet.set_column(15, 15, convert(2))
    worksheet.set_column(16, 16, convert(4.72))
    worksheet.set_column(17, 17, convert(2))
    worksheet.set_column(18, 18, convert(4.72))
    worksheet.set_column(19, 19, convert(2))
    worksheet.set_column(20, 20, convert(4.72))
    worksheet.set_column(21, 21, convert(2))
    worksheet.set_column(22, 22, convert(2.24))
    worksheet.set_column(23, 23, convert(3.36))
    writer.close()
    wb = load_workbook(filename=filename)
    worksheet = wb.active
    rows = ["Название Магазина", "Дата регистрации", "Рейтинг", "Телефон",
            "Количество отзывов", "Успешных заказов"]
    worksheet.insert_rows(0, 9)
    for i, row in enumerate(rows, start=2):
        worksheet[f"B{i}"] = row
        worksheet[f"C{i}"] = merch[row]
    wb.save(filename=filename)
    with zipfile.ZipFile(filename.replace("xlsx", "zip"), mode="w") as archive:
        for dirname, subdirs, files in os.walk(filename.replace(".xlsx", "")):
            archive.write(dirname)
            for filename_dir in files:
                archive.write(os.path.join(dirname, filename_dir))
        archive.write(filename)
        filename = filename.replace("xlsx", "zip")
    wss.put(
        json.dumps(
            {
                "type": mode,
                "curr": len(results),
                "total": len(results),
                "eta": "-:-",
                "id": ind,
                "status": "Готово",
                "name": query,
                "city": cities[city],
                "filename": filename
            }
        )
    )
